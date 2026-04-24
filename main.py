"""
Trivy Docker Image Scanner - FastAPI backend.

Endpoints:
  GET  /                       -> serves the single-page UI

  Trivy:
  GET  /api/trivy/version      -> currently installed Trivy version
  GET  /api/trivy/releases     -> list Trivy releases from GitHub
  POST /api/trivy/install      -> download & install a specific Trivy version

  Docker:
  GET  /api/docker/check       -> is the docker CLI available?
  POST /api/docker/login       -> docker login <registry>
  POST /api/docker/pull        -> docker pull <image> + docker save -> .tar
  GET  /api/docker/tars        -> list .tar files in the output directory

  Scan:
  POST /api/scan               -> scan an uploaded .tar
  POST /api/scan_path          -> scan a .tar already on disk (e.g. just pulled)
"""

from __future__ import annotations

import io
import json
import os
import platform
import re
import shutil
import stat
import subprocess
import sys
import tarfile
import tempfile
import zipfile
from pathlib import Path
from typing import Any

import httpx
from fastapi import Body, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response, StreamingResponse
from pydantic import BaseModel

APP_DIR = Path(__file__).parent.resolve()
BIN_DIR = APP_DIR / "bin"
BIN_DIR.mkdir(exist_ok=True)
TARS_DIR = APP_DIR / "tars"
TARS_DIR.mkdir(exist_ok=True)
TRIVY_BIN = BIN_DIR / ("trivy.exe" if os.name == "nt" else "trivy")
INDEX_HTML = APP_DIR / "index.html"

GITHUB_RELEASES_API = "https://api.github.com/repos/aquasecurity/trivy/releases"
GITHUB_DOWNLOAD_BASE = "https://github.com/aquasecurity/trivy/releases/download"

app = FastAPI(title="Trivy Docker Image Scanner", version="1.1.0")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _trivy_path() -> str | None:
    if TRIVY_BIN.exists():
        return str(TRIVY_BIN)
    return shutil.which("trivy")


def _docker_path() -> str | None:
    """Locate the docker CLI. Falls back to common Docker Desktop install paths
    on Windows so the app still works from a terminal opened before Docker was
    added to PATH."""
    found = shutil.which("docker")
    if found:
        return found
    if os.name == "nt":
        candidates = [
            r"C:\Program Files\Docker\Docker\resources\bin\docker.exe",
            r"C:\Program Files\Docker\Docker\resources\docker.exe",
            os.path.expandvars(r"%ProgramFiles%\Docker\Docker\resources\bin\docker.exe"),
            os.path.expandvars(r"%LOCALAPPDATA%\Docker\Docker\resources\bin\docker.exe"),
        ]
    else:
        candidates = ["/usr/local/bin/docker", "/usr/bin/docker", "/opt/homebrew/bin/docker"]
    for c in candidates:
        if c and Path(c).exists():
            return c
    return None


def _detect_platform() -> tuple[str, str, str]:
    system = platform.system()
    machine = platform.machine().lower()
    if system == "Linux":
        os_name, ext = "Linux", "tar.gz"
    elif system == "Darwin":
        os_name, ext = "macOS", "tar.gz"
    elif system == "Windows":
        os_name, ext = "windows", "zip"
    else:
        raise RuntimeError(f"Unsupported OS: {system}")
    if machine in ("x86_64", "amd64"):
        arch = "64bit"
    elif machine in ("aarch64", "arm64"):
        arch = "ARM64"
    elif machine in ("i386", "i686"):
        arch = "32bit"
    else:
        raise RuntimeError(f"Unsupported arch: {machine}")
    return os_name, arch, ext


def _normalise_version(v: str) -> str:
    return v[1:] if v.startswith("v") else v


def _safe_name_for_image(image: str) -> str:
    """Turn 'registry/.../foo:tag' into 'foo.tar'."""
    last = image.rsplit("/", 1)[-1]
    if ":" in last:
        last = last.split(":", 1)[0]
    last = re.sub(r"[^A-Za-z0-9_.-]", "_", last) or "image"
    return f"{last}.tar"


def _parse_trivy_report(stdout: str) -> dict[str, Any]:
    """Convert raw Trivy JSON output into the shape consumed by the UI."""
    try:
        report = json.loads(stdout)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(500, f"Could not parse Trivy output: {e}")

    summary = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0, "UNKNOWN": 0}
    vulns: list[dict[str, Any]] = []
    for result in report.get("Results", []) or []:
        target = result.get("Target")
        rtype = result.get("Type")
        for v in result.get("Vulnerabilities") or []:
            sev = (v.get("Severity") or "UNKNOWN").upper()
            summary[sev] = summary.get(sev, 0) + 1
            vulns.append({
                "id": v.get("VulnerabilityID"),
                "pkg": v.get("PkgName"),
                "installed": v.get("InstalledVersion"),
                "fixed": v.get("FixedVersion"),
                "severity": sev,
                "title": v.get("Title"),
                "description": v.get("Description"),
                "primary_url": v.get("PrimaryURL"),
                "references": v.get("References") or [],
                "cvss": v.get("CVSS"),
                "target": target,
                "type": rtype,
            })

    meta = {
        "artifact": report.get("ArtifactName"),
        "artifact_type": report.get("ArtifactType"),
        "schema_version": report.get("SchemaVersion"),
        "created_at": report.get("CreatedAt"),
    }
    return {
        "summary": summary,
        "total": sum(summary.values()),
        "vulnerabilities": vulns,
        "meta": meta,
    }


def _run_trivy_scan(tar_path: Path, severity: str, ignore_unfixed: bool) -> dict[str, Any]:
    trivy = _trivy_path()
    if not trivy:
        raise HTTPException(
            400,
            "Trivy is not installed. Install a Trivy version from the right-hand panel first.",
        )
    cmd = [
        trivy, "image",
        "--input", str(tar_path),
        "--format", "json",
        "--severity", severity,
        "--scanners", "vuln",
        "--quiet",
    ]
    if ignore_unfixed:
        cmd.append("--ignore-unfixed")

    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=900)
    if proc.returncode != 0 and not (proc.stdout or "").strip():
        raise HTTPException(
            500, f"Trivy scan failed (exit {proc.returncode}): {proc.stderr[:1000]}"
        )
    return _parse_trivy_report(proc.stdout)


def _run_trivy_scan_image(image: str, severity: str, ignore_unfixed: bool) -> dict[str, Any]:
    """Quick scan: hand the image name to Trivy directly. Trivy uses the local
    Docker daemon (or pulls the image from the registry) — no .tar is written."""
    trivy = _trivy_path()
    if not trivy:
        raise HTTPException(
            400,
            "Trivy is not installed. Install a Trivy version from the right-hand panel first.",
        )
    cmd = [
        trivy, "image",
        "--format", "json",
        "--severity", severity,
        "--scanners", "vuln",
        "--quiet",
        image,
    ]
    if ignore_unfixed:
        cmd.append("--ignore-unfixed")

    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=1800)
    if proc.returncode != 0 and not (proc.stdout or "").strip():
        raise HTTPException(
            500,
            f"Trivy scan failed for {image} (exit {proc.returncode}): "
            f"{(proc.stderr or proc.stdout).strip()[:1000]}",
        )
    return _parse_trivy_report(proc.stdout)


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------
@app.get("/")
async def root():
    if not INDEX_HTML.exists():
        raise HTTPException(500, "index.html not found next to main.py")
    # Disable browser caching so UI changes show up immediately on reload.
    return FileResponse(
        INDEX_HTML,
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


# ---------------------------------------------------------------------------
# Trivy
# ---------------------------------------------------------------------------
@app.get("/api/trivy/version")
async def trivy_version() -> dict[str, Any]:
    path = _trivy_path()
    if not path:
        return {"installed": False, "version": None, "path": None}
    try:
        out = subprocess.run([path, "--version"], capture_output=True, text=True, timeout=15)
        first = (out.stdout or out.stderr or "").strip().splitlines()
        version = None
        for line in first:
            m = re.search(r"Version:\s*([0-9][^\s]+)", line)
            if m:
                version = m.group(1)
                break
        if not version and first:
            version = first[0]
        return {"installed": True, "version": version, "path": path}
    except Exception as e:  # noqa: BLE001
        return {"installed": False, "version": None, "path": path, "error": str(e)}


@app.get("/api/trivy/releases")
async def trivy_releases(limit: int = 30) -> dict[str, Any]:
    async with httpx.AsyncClient(timeout=30.0) as client:
        r = await client.get(
            GITHUB_RELEASES_API,
            params={"per_page": min(max(limit, 1), 100)},
            headers={"Accept": "application/vnd.github+json"},
        )
    if r.status_code != 200:
        raise HTTPException(502, f"GitHub API error: {r.status_code} {r.text[:200]}")
    return {
        "releases": [
            {
                "tag": rel.get("tag_name"),
                "name": rel.get("name"),
                "published_at": rel.get("published_at"),
                "prerelease": rel.get("prerelease", False),
                "html_url": rel.get("html_url"),
            }
            for rel in r.json()
        ]
    }


@app.post("/api/trivy/install")
async def trivy_install(version: str = Form(...)) -> dict[str, Any]:
    ver = _normalise_version(version.strip())
    if not re.match(r"^\d+\.\d+\.\d+(-\S+)?$", ver):
        raise HTTPException(400, f"Invalid version: {version}")
    try:
        os_name, arch, ext = _detect_platform()
    except RuntimeError as e:
        raise HTTPException(500, str(e))

    asset = f"trivy_{ver}_{os_name}-{arch}.{ext}"
    url = f"{GITHUB_DOWNLOAD_BASE}/v{ver}/{asset}"
    async with httpx.AsyncClient(timeout=120.0, follow_redirects=True) as client:
        r = await client.get(url)
    if r.status_code != 200:
        raise HTTPException(
            404,
            f"Failed to download asset {asset} (HTTP {r.status_code}). "
            f"Check that this version exists for your platform.",
        )

    raw = io.BytesIO(r.content)
    tmpdir = Path(tempfile.mkdtemp(prefix="trivy_install_"))
    try:
        if ext == "zip":
            with zipfile.ZipFile(raw) as zf:
                zf.extractall(tmpdir)
        else:
            with tarfile.open(fileobj=raw, mode="r:gz") as tf:
                tf.extractall(tmpdir)

        candidate = None
        for p in tmpdir.rglob("trivy*"):
            if p.is_file() and p.name.lower().startswith("trivy"):
                if p.suffix.lower() in (".exe", "") and "LICENSE" not in p.name.upper():
                    candidate = p
                    break
        if not candidate:
            raise HTTPException(500, "Trivy binary not found in downloaded archive")

        if TRIVY_BIN.exists():
            TRIVY_BIN.unlink()
        shutil.copy2(candidate, TRIVY_BIN)
        if os.name != "nt":
            TRIVY_BIN.chmod(
                TRIVY_BIN.stat().st_mode | stat.S_IEXEC | stat.S_IXGRP | stat.S_IXOTH
            )
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    return {"installed": True, "version": ver, "path": str(TRIVY_BIN), "asset": asset}


# ---------------------------------------------------------------------------
# Docker
# ---------------------------------------------------------------------------
@app.get("/api/docker/check")
async def docker_check() -> dict[str, Any]:
    docker = _docker_path()
    if not docker:
        return {"available": False, "default_dir": str(TARS_DIR)}
    try:
        out = subprocess.run(
            [docker, "version", "--format", "{{.Client.Version}}"],
            capture_output=True, text=True, timeout=15,
        )
        ver = (out.stdout or "").strip()
        return {
            "available": out.returncode == 0 and bool(ver),
            "version": ver,
            "path": docker,
            "default_dir": str(TARS_DIR),
        }
    except Exception as e:  # noqa: BLE001
        return {
            "available": False,
            "error": str(e),
            "path": docker,
            "default_dir": str(TARS_DIR),
        }


@app.post("/api/docker/login")
async def docker_login(
    registry: str = Form(...),
    username: str = Form(...),
    password: str = Form(...),
) -> dict[str, Any]:
    docker = _docker_path()
    if not docker:
        raise HTTPException(400, "Docker is not installed on the host.")
    try:
        proc = subprocess.run(
            [docker, "login", "-u", username, "--password-stdin", registry],
            input=password, capture_output=True, text=True, timeout=60,
        )
    except subprocess.TimeoutExpired:
        raise HTTPException(
            408,
            "Docker login timed out after 60s — check the registry hostname / network.",
        )
    if proc.returncode != 0:
        err = (proc.stderr or proc.stdout or "").strip()
        # Surface the actual error verbatim so the frontend can interpret it.
        raise HTTPException(
            400,
            err[:1500] if err else f"docker login exited with code {proc.returncode}",
        )
    return {
        "ok": True,
        "registry": registry,
        "username": username,
        "message": (proc.stdout or "Login Succeeded").strip(),
    }


@app.post("/api/docker/logout")
async def docker_logout(registry: str = Form(...)) -> dict[str, Any]:
    docker = _docker_path()
    if not docker:
        raise HTTPException(400, "Docker is not installed on the host.")
    proc = subprocess.run(
        [docker, "logout", registry],
        capture_output=True, text=True, timeout=30,
    )
    if proc.returncode != 0:
        raise HTTPException(
            400,
            f"Logout failed: {(proc.stderr or proc.stdout).strip()[:500]}",
        )
    return {
        "ok": True,
        "registry": registry,
        "message": (proc.stdout or "Logout Succeeded").strip(),
    }


class PullRequest(BaseModel):
    images: list[str]
    output_dir: str | None = None


@app.post("/api/docker/pull")
async def docker_pull(req: PullRequest) -> dict[str, Any]:
    docker = _docker_path()
    if not docker:
        raise HTTPException(400, "Docker is not installed on the host.")

    out_dir = (
        Path(req.output_dir).expanduser().resolve()
        if req.output_dir and req.output_dir.strip()
        else TARS_DIR
    )
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        raise HTTPException(400, f"Cannot create output directory '{out_dir}': {e}")

    results: list[dict[str, Any]] = []
    for raw in req.images:
        image = (raw or "").strip()
        if not image:
            continue

        tar_name = _safe_name_for_image(image)
        tar_path = out_dir / tar_name

        # docker pull
        pull = subprocess.run(
            [docker, "pull", image], capture_output=True, text=True, timeout=1800
        )
        if pull.returncode != 0:
            results.append({
                "image": image, "ok": False, "stage": "pull",
                "error": (pull.stderr or pull.stdout).strip()[:1500],
            })
            continue

        # docker save
        save = subprocess.run(
            [docker, "save", "-o", str(tar_path), image],
            capture_output=True, text=True, timeout=1800,
        )
        if save.returncode != 0:
            results.append({
                "image": image, "ok": False, "stage": "save",
                "error": (save.stderr or save.stdout).strip()[:1500],
            })
            continue

        size = tar_path.stat().st_size if tar_path.exists() else 0
        results.append({
            "image": image, "ok": True,
            "tar_path": str(tar_path),
            "tar_name": tar_path.name,
            "size_bytes": size,
        })

    return {"results": results, "output_dir": str(out_dir)}


def _ndjson(event: str, data: Any) -> bytes:
    """Encode one newline-delimited JSON event for streaming responses."""
    return (json.dumps({"event": event, "data": data}) + "\n").encode("utf-8")


@app.post("/api/docker/pull_stream")
def docker_pull_stream(req: PullRequest) -> StreamingResponse:
    """Streams `docker pull` + `docker save` output line-by-line as NDJSON
    so the UI can show a live log instead of a blocking spinner."""
    docker = _docker_path()
    if not docker:
        raise HTTPException(400, "Docker is not installed on the host.")

    out_dir = (
        Path(req.output_dir).expanduser().resolve()
        if req.output_dir and req.output_dir.strip()
        else TARS_DIR
    )
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        raise HTTPException(400, f"Cannot create output directory '{out_dir}': {e}")

    images = [(i or "").strip() for i in req.images]
    images = [i for i in images if i]
    if not images:
        raise HTTPException(400, "No image URLs provided.")

    def stream():
        yield _ndjson("init", {"output_dir": str(out_dir), "total": len(images)})
        for idx, image in enumerate(images, start=1):
            tar_name = _safe_name_for_image(image)
            tar_path = out_dir / tar_name
            yield _ndjson("image_start", {
                "image": image, "index": idx, "total": len(images),
                "tar_path": str(tar_path),
            })

            # ---- docker pull ----
            yield _ndjson("log", {"line": f"$ docker pull {image}"})
            try:
                proc = subprocess.Popen(
                    [docker, "pull", image],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=1,
                )
                for line in iter(proc.stdout.readline, ""):
                    yield _ndjson("log", {"line": line.rstrip()})
                proc.wait(timeout=1800)
            except subprocess.TimeoutExpired:
                proc.kill()
                yield _ndjson("image_done", {
                    "image": image, "ok": False, "stage": "pull",
                    "error": "docker pull timed out after 30 minutes",
                })
                continue
            except Exception as e:  # noqa: BLE001
                yield _ndjson("image_done", {
                    "image": image, "ok": False, "stage": "pull", "error": str(e),
                })
                continue
            if proc.returncode != 0:
                yield _ndjson("image_done", {
                    "image": image, "ok": False, "stage": "pull",
                    "error": f"docker pull exited with code {proc.returncode}",
                })
                continue

            # ---- docker save ----
            yield _ndjson("log", {"line": f"$ docker save -o {tar_path} {image}"})
            try:
                save = subprocess.Popen(
                    [docker, "save", "-o", str(tar_path), image],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=1,
                )
                for line in iter(save.stdout.readline, ""):
                    yield _ndjson("log", {"line": line.rstrip()})
                save.wait(timeout=1800)
            except subprocess.TimeoutExpired:
                save.kill()
                yield _ndjson("image_done", {
                    "image": image, "ok": False, "stage": "save",
                    "error": "docker save timed out after 30 minutes",
                })
                continue
            except Exception as e:  # noqa: BLE001
                yield _ndjson("image_done", {
                    "image": image, "ok": False, "stage": "save", "error": str(e),
                })
                continue
            if save.returncode != 0:
                yield _ndjson("image_done", {
                    "image": image, "ok": False, "stage": "save",
                    "error": f"docker save exited with code {save.returncode}",
                })
                continue

            size = tar_path.stat().st_size if tar_path.exists() else 0
            human = (
                f"{size:,} B" if size < 1024 else
                f"{size/1024:.1f} KB" if size < 1024 * 1024 else
                f"{size/1024/1024:.1f} MB" if size < 1024 ** 3 else
                f"{size/1024/1024/1024:.2f} GB"
            )
            yield _ndjson("log", {"line": f"OK saved {tar_path.name} ({human})"})
            yield _ndjson("image_done", {
                "image": image, "ok": True,
                "tar_path": str(tar_path),
                "tar_name": tar_path.name,
                "size_bytes": size,
            })
        yield _ndjson("done", {"output_dir": str(out_dir)})

    return StreamingResponse(
        stream(),
        media_type="application/x-ndjson",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/docker/tars")
async def list_tars(output_dir: str | None = None) -> dict[str, Any]:
    out_dir = (
        Path(output_dir).expanduser().resolve()
        if output_dir and output_dir.strip()
        else TARS_DIR
    )
    if not out_dir.exists():
        return {"output_dir": str(out_dir), "tars": []}
    items = []
    for p in sorted(out_dir.glob("*.tar"), key=lambda x: -x.stat().st_mtime):
        st = p.stat()
        items.append({
            "name": p.name,
            "path": str(p),
            "size_bytes": st.st_size,
            "modified": st.st_mtime,
        })
    return {"output_dir": str(out_dir), "tars": items}


# ---------------------------------------------------------------------------
# Folder picker (native OS dialog — only works because this app runs locally)
# ---------------------------------------------------------------------------
@app.post("/api/browse_folder")
async def browse_folder(initial_dir: str = Form("")) -> dict[str, Any]:
    """Open a native folder picker on the host machine and return the chosen path."""
    init = initial_dir.strip() or str(TARS_DIR)
    code = (
        "import tkinter as tk\n"
        "from tkinter import filedialog\n"
        "root = tk.Tk()\n"
        "root.withdraw()\n"
        "root.attributes('-topmost', True)\n"
        f"folder = filedialog.askdirectory(title='Select output folder for .tar files', initialdir=r'{init}')\n"
        "root.destroy()\n"
        "print(folder or '')\n"
    )
    try:
        proc = subprocess.run(
            [sys.executable, "-c", code],
            capture_output=True,
            text=True,
            timeout=300,
        )
    except subprocess.TimeoutExpired:
        raise HTTPException(408, "Folder picker timed out (no folder chosen).")
    if proc.returncode != 0:
        raise HTTPException(
            500,
            f"Folder picker failed: {(proc.stderr or proc.stdout).strip()[:500]}",
        )
    folder = (proc.stdout or "").strip().splitlines()[-1] if proc.stdout.strip() else ""
    if not folder:
        return {"folder": None, "cancelled": True}
    return {"folder": str(Path(folder).expanduser().resolve()), "cancelled": False}


# ---------------------------------------------------------------------------
# Scan
# ---------------------------------------------------------------------------
@app.post("/api/scan")
async def scan_image(
    file: UploadFile = File(...),
    severity: str = Form("CRITICAL,HIGH,MEDIUM,LOW,UNKNOWN"),
    ignore_unfixed: bool = Form(False),
) -> JSONResponse:
    tmp_dir = Path(tempfile.mkdtemp(prefix="trivy_scan_"))
    tar_path = tmp_dir / (file.filename or "image.tar")
    try:
        with tar_path.open("wb") as f:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                f.write(chunk)
        return JSONResponse(_run_trivy_scan(tar_path, severity, ignore_unfixed))
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@app.post("/api/scan_path")
async def scan_path(
    path: str = Form(...),
    severity: str = Form("CRITICAL,HIGH,MEDIUM,LOW,UNKNOWN"),
    ignore_unfixed: bool = Form(False),
) -> JSONResponse:
    p = Path(path).expanduser().resolve()
    if not p.exists() or not p.is_file():
        raise HTTPException(404, f"File not found: {p}")
    return JSONResponse(_run_trivy_scan(p, severity, ignore_unfixed))


@app.post("/api/scan_image")
async def scan_image_direct(
    image: str = Form(...),
    severity: str = Form("CRITICAL,HIGH,MEDIUM,LOW,UNKNOWN"),
    ignore_unfixed: bool = Form(False),
) -> JSONResponse:
    """Quick scan: scan a Docker image by name without saving a .tar file."""
    img = (image or "").strip()
    if not img:
        raise HTTPException(400, "Image name is required.")
    return JSONResponse(_run_trivy_scan_image(img, severity, ignore_unfixed))


def _stream_trivy(cmd: list[str], timeout: int = 1800):
    """Generic helper: run a `trivy ... --format json` command, stream stderr
    line-by-line, capture stdout to a temp file, then yield the parsed result."""
    yield _ndjson("log", {"line": f"$ {' '.join(cmd)}"})
    yield _ndjson("log", {"line": "Starting Trivy… (first run downloads the vulnerability DB which can take 30-60s with no output)"})
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as tf:
        out_path = tf.name
    try:
        with open(out_path, "w", encoding="utf-8") as out_file:
            proc = subprocess.Popen(
                cmd,
                stdout=out_file,
                stderr=subprocess.PIPE,
                text=True,
                bufsize=1,
            )
            try:
                for line in iter(proc.stderr.readline, ""):
                    line = line.rstrip()
                    if line:
                        yield _ndjson("log", {"line": line})
                proc.wait(timeout=timeout)
            except subprocess.TimeoutExpired:
                proc.kill()
                yield _ndjson("error", {"message": f"Trivy scan timed out after {timeout}s."})
                return

        with open(out_path, "r", encoding="utf-8") as f:
            stdout_text = f.read()
        if proc.returncode != 0 and not stdout_text.strip():
            yield _ndjson("error", {"message": f"Trivy scan failed (exit {proc.returncode})."})
            return
        try:
            result = _parse_trivy_report(stdout_text)
        except HTTPException as he:
            yield _ndjson("error", {"message": he.detail})
            return
        except Exception as e:  # noqa: BLE001
            yield _ndjson("error", {"message": str(e)})
            return
        yield _ndjson("log", {"line": f"OK scan complete · {result['total']} findings"})
        yield _ndjson("result", result)
        yield _ndjson("done", {})
    finally:
        try: os.unlink(out_path)
        except Exception: pass


@app.post("/api/scan_path_stream")
def scan_path_stream(
    path: str = Form(...),
    severity: str = Form("CRITICAL,HIGH,MEDIUM,LOW,UNKNOWN"),
    ignore_unfixed: bool = Form(False),
) -> StreamingResponse:
    """Stream a Trivy scan of a tar file already on disk."""
    trivy = _trivy_path()
    if not trivy:
        raise HTTPException(
            400,
            "Trivy is not installed. Install a Trivy version from the right-hand panel first.",
        )
    p = Path(path).expanduser().resolve()
    if not p.exists() or not p.is_file():
        raise HTTPException(404, f"File not found: {p}")
    cmd = [
        trivy, "image",
        "--input", str(p),
        "--format", "json",
        "--severity", severity,
        "--scanners", "vuln",
        # NOTE: no --quiet flag — we WANT Trivy's INFO progress output to
        # stream into the live log so the user can see what's happening.
    ]
    if ignore_unfixed:
        cmd.append("--ignore-unfixed")
    return StreamingResponse(
        _stream_trivy(cmd, timeout=900),
        media_type="application/x-ndjson",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/scan_image_stream")
def scan_image_stream(
    image: str = Form(...),
    severity: str = Form("CRITICAL,HIGH,MEDIUM,LOW,UNKNOWN"),
    ignore_unfixed: bool = Form(False),
) -> StreamingResponse:
    """Quick scan with live log streaming. Trivy writes its progress to stderr
    (downloading DB, scanning layers, …) — we stream those lines back, then send
    the parsed result as a final 'result' event."""
    trivy = _trivy_path()
    if not trivy:
        raise HTTPException(
            400,
            "Trivy is not installed. Install a Trivy version from the right-hand panel first.",
        )
    img = (image or "").strip()
    if not img:
        raise HTTPException(400, "Image name is required.")

    cmd = [
        trivy, "image",
        "--format", "json",
        "--severity", severity,
        "--scanners", "vuln",
        img,
    ]
    if ignore_unfixed:
        cmd.append("--ignore-unfixed")
    return StreamingResponse(
        _stream_trivy(cmd, timeout=1800),
        media_type="application/x-ndjson",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ---------------------------------------------------------------------------
# Export (Excel / PDF)
# ---------------------------------------------------------------------------
import io
from datetime import datetime

_SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "UNKNOWN": 4}
_SEV_COLOR_HEX = {  # background fills for the Severity column / badges
    "CRITICAL": "FECACA",
    "HIGH":     "FED7D7",
    "MEDIUM":   "FDE68A",
    "LOW":      "BFDBFE",
    "UNKNOWN":  "E2E8F0",
}


def _sorted_vulns(vulns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        vulns,
        key=lambda v: (
            _SEV_ORDER.get((v.get("severity") or "UNKNOWN").upper(), 9),
            (v.get("pkg") or "").lower(),
            (v.get("id") or "").lower(),
        ),
    )


def _safe_filename(stem: str) -> str:
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem or "trivy-report").strip("._-")
    return stem or "trivy-report"


@app.post("/api/export/xlsx")
async def export_xlsx(payload: dict[str, Any] = Body(...)) -> Response:
    """Export the scan result as a multi-sheet .xlsx workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            500,
            "openpyxl is not installed on the backend. Run `pip install -r requirements.txt`.",
        )

    summary = payload.get("summary") or {}
    vulns = _sorted_vulns(payload.get("vulnerabilities") or [])
    meta = payload.get("meta") or {}
    artifact = meta.get("artifact") or "image"

    wb = Workbook()

    # --- Summary sheet ---
    s = wb.active
    s.title = "Summary"
    title = s.cell(1, 1, "Trivy Scan Report")
    title.font = Font(size=16, bold=True)
    s.cell(2, 1, "Artifact:")
    s.cell(2, 2, artifact)
    s.cell(3, 1, "Type:")
    s.cell(3, 2, meta.get("artifact_type") or "")
    s.cell(4, 1, "Generated:")
    s.cell(4, 2, datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"))
    s.cell(5, 1, "Total findings:")
    s.cell(5, 2, payload.get("total") or 0)

    s.cell(7, 1, "Severity").font = Font(bold=True)
    s.cell(7, 2, "Count").font = Font(bold=True)
    for i, sev in enumerate(["CRITICAL", "HIGH", "MEDIUM", "LOW", "UNKNOWN"], start=8):
        c = s.cell(i, 1, sev)
        c.fill = PatternFill("solid", fgColor=_SEV_COLOR_HEX.get(sev, "FFFFFF"))
        c.font = Font(bold=True)
        s.cell(i, 2, summary.get(sev, 0))
    s.column_dimensions["A"].width = 18
    s.column_dimensions["B"].width = 60

    # --- Vulnerabilities sheet ---
    v = wb.create_sheet("Vulnerabilities")
    headers = ["CVE", "Package", "Installed", "Fixed in", "Severity",
               "Title", "Target", "Type", "Primary URL"]
    for col, h in enumerate(headers, start=1):
        c = v.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="334155")
        c.alignment = Alignment(vertical="center")
    widths = [22, 28, 22, 22, 12, 60, 30, 14, 50]
    for col, w in enumerate(widths, start=1):
        v.column_dimensions[get_column_letter(col)].width = w
    v.freeze_panes = "A2"

    for i, vuln in enumerate(vulns, start=2):
        sev = (vuln.get("severity") or "UNKNOWN").upper()
        v.cell(i, 1, vuln.get("id"))
        v.cell(i, 2, vuln.get("pkg"))
        v.cell(i, 3, vuln.get("installed"))
        v.cell(i, 4, vuln.get("fixed"))
        cell_sev = v.cell(i, 5, sev)
        cell_sev.fill = PatternFill("solid", fgColor=_SEV_COLOR_HEX.get(sev, "FFFFFF"))
        cell_sev.font = Font(bold=True)
        v.cell(i, 6, vuln.get("title"))
        v.cell(i, 7, vuln.get("target"))
        v.cell(i, 8, vuln.get("type"))
        v.cell(i, 9, vuln.get("primary_url"))

    if v.auto_filter:
        v.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(2, len(vulns)+1)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = _safe_filename(artifact.rsplit("/", 1)[-1].split(":")[0]) + "_trivy.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


@app.post("/api/export/pdf")
async def export_pdf(payload: dict[str, Any] = Body(...)) -> Response:
    """Export the scan result as a printable PDF."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
        )
    except ImportError:
        raise HTTPException(
            500,
            "reportlab is not installed on the backend. Run `pip install -r requirements.txt`.",
        )

    summary = payload.get("summary") or {}
    vulns = _sorted_vulns(payload.get("vulnerabilities") or [])
    meta = payload.get("meta") or {}
    artifact = meta.get("artifact") or "image"

    sev_colors = {
        "CRITICAL": colors.HexColor("#fee2e2"),
        "HIGH":     colors.HexColor("#fed7d7"),
        "MEDIUM":   colors.HexColor("#fde68a"),
        "LOW":      colors.HexColor("#bfdbfe"),
        "UNKNOWN":  colors.HexColor("#e2e8f0"),
    }

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=f"Trivy Scan Report — {artifact}",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=16, spaceAfter=4)
    meta_style = ParagraphStyle("Meta", parent=styles["BodyText"], fontSize=9, textColor=colors.HexColor("#475569"))
    cell_style = ParagraphStyle("Cell", parent=styles["BodyText"], fontSize=8, leading=10)
    cell_mono = ParagraphStyle("CellMono", parent=cell_style, fontName="Courier", fontSize=7.5)

    story = []
    story.append(Paragraph("Trivy Scan Report", h1))
    story.append(Paragraph(f"<b>Artifact:</b> {artifact}", meta_style))
    if meta.get("artifact_type"):
        story.append(Paragraph(f"<b>Type:</b> {meta.get('artifact_type')}", meta_style))
    story.append(Paragraph(f"<b>Generated:</b> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}", meta_style))
    story.append(Paragraph(f"<b>Total findings:</b> {payload.get('total') or 0}", meta_style))
    story.append(Spacer(1, 6 * mm))

    # Severity summary
    summary_data = [["CRITICAL", "HIGH", "MEDIUM", "LOW", "UNKNOWN"],
                    [summary.get(k, 0) for k in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "UNKNOWN"]]]
    sum_table = Table(summary_data, colWidths=[40 * mm] * 5)
    sum_style = TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, 1), 18),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
    ])
    for i, sev in enumerate(["CRITICAL", "HIGH", "MEDIUM", "LOW", "UNKNOWN"]):
        sum_style.add("BACKGROUND", (i, 0), (i, 0), sev_colors[sev])
    sum_table.setStyle(sum_style)
    story.append(sum_table)
    story.append(Spacer(1, 6 * mm))

    # Vulnerabilities table
    headers = ["CVE", "Package", "Installed", "Fixed in", "Sev", "Title"]
    col_widths = [38 * mm, 38 * mm, 28 * mm, 28 * mm, 18 * mm, 120 * mm]
    body = [headers]
    for v in vulns:
        body.append([
            Paragraph(v.get("id") or "", cell_mono),
            Paragraph(v.get("pkg") or "", cell_mono),
            Paragraph(v.get("installed") or "", cell_mono),
            Paragraph(v.get("fixed") or "—", cell_mono),
            (v.get("severity") or "UNKNOWN").upper(),
            Paragraph(v.get("title") or "", cell_style),
        ])

    if len(body) == 1:
        body.append(["", "", "", "", "", "No vulnerabilities found."])

    t = Table(body, colWidths=col_widths, repeatRows=1)
    ts = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (4, 1), (4, -1), "CENTER"),
        ("FONTNAME", (4, 1), (4, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
    ])
    for i, v in enumerate(vulns, start=1):
        sev = (v.get("severity") or "UNKNOWN").upper()
        ts.add("BACKGROUND", (4, i), (4, i), sev_colors.get(sev, colors.white))
    t.setStyle(ts)
    story.append(t)

    doc.build(story)
    buf.seek(0)
    fn = _safe_filename(artifact.rsplit("/", 1)[-1].split(":")[0]) + "_trivy.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", "8000"))
    url = f"http://localhost:{port}"
    print("=" * 60)
    print(f"  Trivy Docker Image Scanner")
    print(f"  Running on: {url}")
    print("=" * 60)
    host = os.environ.get("HOST", "localhost")
    uvicorn.run(
        "main:app",
        host=host,
        port=port,
        reload=False,
        forwarded_allow_ips="*",
    )
